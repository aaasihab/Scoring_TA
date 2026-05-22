import os
from datetime import datetime
from openpyxl import Workbook, load_workbook


HISTORY_FILE = os.path.join(os.path.dirname(os.path.dirname(__file__)), "dataset", "history.xlsx")

HEADERS = [
    "No", "Timestamp", "Mode", "Query (Raw)", "Query (Preprocessed)",
    "Unique Score (Max)", "Label Unique (Max)",
    "Top 1 - Text", "Top 1 - Score", "Top 1 - Label",
    "Top 2 - Text", "Top 2 - Score", "Top 2 - Label",
    "Top 3 - Text", "Top 3 - Score", "Top 3 - Label",
    "Top 4 - Text", "Top 4 - Score", "Top 4 - Label",
    "Top 5 - Text", "Top 5 - Score", "Top 5 - Label",
]


# ======================================
# Simpan Hasil ke History Excel
# ======================================

def save_to_history(results):

    if os.path.exists(HISTORY_FILE):
        wb = load_workbook(HISTORY_FILE)
        ws = wb.active
        next_no = ws.max_row  # max_row includes header row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(HEADERS)
        next_no = 1

    row = [
        next_no,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        results["mode"],
        results["raw_query"],
        results["query"],
        results["unique"]["unique_max"],
        results["unique"]["label_max"],
    ]

    # Append Top 1-5 results
    for r in results["results"]:
        row.extend([r["text"], r["score"], r["label"]])

    ws.append(row)
    wb.save(HISTORY_FILE)


# ======================================
# Muat Semua Data History
# ======================================

def load_history():

    if not os.path.exists(HISTORY_FILE):
        return []

    wb = load_workbook(HISTORY_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        data.append(record)

    data.reverse()  # Terbaru di atas
    return data
