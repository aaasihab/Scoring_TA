import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Path file Excel untuk menyimpan history pengujian
HISTORY_FILE = os.path.join(
    os.path.dirname(os.path.dirname(__file__)),
    "dataset",
    "history.xlsx"
)

# Daftar kolom header untuk file Excel history
# Urutan harus sesuai dengan urutan data di fungsi save_to_history()
HEADERS = [
    "No", "Timestamp", "Mode", "Query (Raw)", "Query (Preprocessed)",
    "Unique Score (Max)", "Label Unique (Max)",
    "Top 1 - Text", "Top 1 - Score", "Top 1 - Label",
    "Top 2 - Text", "Top 2 - Score", "Top 2 - Label",
    "Top 3 - Text", "Top 3 - Score", "Top 3 - Label",
    "Top 4 - Text", "Top 4 - Score", "Top 4 - Label",
    "Top 5 - Text", "Top 5 - Score", "Top 5 - Label",
    "Query (Raw Judul)", "Query (Raw Deskripsi)",
    "Top 1 - Judul", "Top 1 - Deskripsi",
    "Top 2 - Judul", "Top 2 - Deskripsi",
    "Top 3 - Judul", "Top 3 - Deskripsi",
    "Top 4 - Judul", "Top 4 - Deskripsi",
    "Top 5 - Judul", "Top 5 - Deskripsi",
]

# Simpan hasil pencarian kemiripan ke file Excel history
def save_to_history(results):

    # Buka file existing atau buat baru jika belum ada
    if os.path.exists(HISTORY_FILE):
        wb = load_workbook(HISTORY_FILE)
        ws = wb.active
        next_no = ws.max_row
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "History"
        ws.append(HEADERS)
        next_no = 1

    # Susun baris data: metadata + skor keunikan
    row = [
        next_no,
        datetime.now(),
        results["mode"],
        results["raw_query"],
        results["query"],
        results["unique"]["unique_max"],
        results["unique"]["label_max"],
    ]

    # Tambahkan Top 1-5 hasil kemiripan (text, score, label)
    for r in results["results"]:
        row.extend([
            r["text"],
            r["score"],
            r["label"]
        ])

    # Tambahkan query judul & deskripsi (khusus mode kombinasi, kosong jika mode lain)
    row.append(results.get("query_judul", ""))
    row.append(results.get("query_deskripsi", ""))

    # Tambahkan Top 1-5 judul & deskripsi terpisah (khusus mode kombinasi)
    for r in results["results"]:
        row.extend([
            r.get("text_judul", ""),
            r.get("text_deskripsi", "")
        ])

    # Tambah data ke worksheet
    ws.append(row)

    # Format kolom Timestamp (kolom B) agar terbaca sebagai tanggal
    timestamp_cell = ws.cell(row=ws.max_row, column=2)
    timestamp_cell.number_format = "dd/mm/yyyy hh:mm"

    # Simpan file
    wb.save(HISTORY_FILE)


# Muat seluruh data history dari file Excel, diurutkan terbaru di atas
def load_history():

    if not os.path.exists(HISTORY_FILE):
        return []

    wb = load_workbook(HISTORY_FILE)
    ws = wb.active

    # Ambil header dari baris pertama sebagai key untuk dictionary
    headers = [cell.value for cell in ws[1]]
    data = []

    # Konversi setiap baris data menjadi dictionary {header: value}
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        # Format Timestamp dari objek datetime ke string yang terbaca
        if isinstance(record.get("Timestamp"), datetime):
            record["Timestamp"] = record["Timestamp"].strftime("%d/%m/%Y %H:%M")
        data.append(record)

    data.reverse()  # Data terbaru di atas
    return data